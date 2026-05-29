"""
Balance Confirmation – Bulk Email App (Streamlit)
==================================================
Run:
    pip install -r requirements.txt
    streamlit run conf_email.py
"""

import streamlit as st
import openpyxl
import smtplib
import ssl
import re
import io
import time
import pandas as pd
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, date

st.set_page_config(page_title="Balance Confirmation Mailer", page_icon="📧",
                   layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=DM+Sans:wght@400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.stTabs [data-baseweb="tab-list"] { gap: 4px; border-bottom: 2px solid var(--primary-color); }
.stTabs [data-baseweb="tab"] { font-family: 'IBM Plex Mono', monospace; font-size: 11px; letter-spacing: 1.5px; text-transform: uppercase; padding: 10px 20px; border-radius: 4px 4px 0 0; }
.sec-head { font-family: 'IBM Plex Mono', monospace; font-size: 10px; letter-spacing: 2px; text-transform: uppercase; opacity: 0.7; border-bottom: 1px solid var(--border-color); padding-bottom: 6px; margin: 16px 0 12px; }
.metric-box { background: var(--secondary-background-color); border: 1px solid var(--border-color); border-radius: 6px; padding: 14px 18px; text-align: center; }
.metric-val { font-size: 24px; font-weight: bold; color: var(--primary-color); }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# PLACEHOLDERS
# ══════════════════════════════════════════════════════════════════════════════

PLACEHOLDERS = {
    "{{Party Name}}":              lambda r, cfg: str(r.get("Party Name", "")),
    "{{Email ID}}":                lambda r, cfg: str(r.get("Email ID", "")),
    "{{Contact Person}}":          lambda r, cfg: str(r.get("Contact Person") or "Sir/Madam"),
    "{{Type (AR/AP)}}":            lambda r, cfg: "AP" if "AP" in str(r.get("Type (AR/AP)", "")).upper() else "AR",
    "{{Currency}}":                lambda r, cfg: str(r.get("Currency") or "INR"),
    "{{Outstanding Balance}}":     lambda r, cfg: fmt_amount(r.get("Outstanding Balance", 0), r.get("Currency") or "INR"),
    "{{Due Date}}":                lambda r, cfg: str(r.get("Due Date", ""))[:10] if r.get("Due Date") else "as agreed",
    "{{Reference / Invoice No.}}": lambda r, cfg: str(r.get("Reference / Invoice No.") or "N/A"),
    "{{Remarks}}":                 lambda r, cfg: str(r.get("Remarks") or ""),
    "{{Confirmation Date}}":       lambda r, cfg: cfg.get("conf_date", ""),
    "{{Company Name}}":            lambda r, cfg: cfg.get("company", ""),
    "{{Signatory}}":               lambda r, cfg: cfg.get("signatory", ""),
    "{{Reply-To Email}}":          lambda r, cfg: cfg.get("reply_to", ""),
    "{{Flow}}":                    lambda r, cfg: "payable to your organisation" if "AP" in str(r.get("Type (AR/AP)", "")).upper() else "receivable from your organisation",
}

DEFAULT_SUBJECT = "Balance Confirmation as on {{Confirmation Date}} – {{Party Name}} [Ref: {{Reference / Invoice No.}}]"

DEFAULT_BODY = """Dear {{Contact Person}},

Greetings from {{Company Name}}!

As part of our periodic balance confirmation exercise, we kindly request you to confirm the outstanding balance as on {{Confirmation Date}}.

As per our records, the following amount is {{Flow}}:

Party Name          : {{Party Name}}
Outstanding Amount  : {{Outstanding Balance}}
Reference           : {{Reference / Invoice No.}}
Due Date            : {{Due Date}}
Remarks             : {{Remarks}}

Kindly confirm the above balance by replying to this email. If there is any discrepancy, please share the relevant details and supporting documents so that we may reconcile at the earliest.

Your prompt response will be greatly appreciated.

For any queries, please write to us at: {{Reply-To Email}}

Warm regards,
{{Signatory}}
{{Company Name}}"""

# ══════════════════════════════════════════════════════════════════════════════
# EMAIL STYLES
# ══════════════════════════════════════════════════════════════════════════════

EMAIL_STYLES = {
    "Professional Blue": {
        "font":        "Arial, Helvetica, sans-serif",
        "header_bg":   "#1F4E79",
        "header_text": "#FFFFFF",
        "table_header_bg": "#2E75B6",
        "table_header_text": "#FFFFFF",
        "table_row_bg":  "#EBF3FB",
        "table_alt_bg":  "#FFFFFF",
        "accent":      "#2E75B6",
        "border":      "#BDD7EE",
    },
    "Classic Black": {
        "font":        "Georgia, 'Times New Roman', serif",
        "header_bg":   "#1A1A1A",
        "header_text": "#FFFFFF",
        "table_header_bg": "#333333",
        "table_header_text": "#FFFFFF",
        "table_row_bg":  "#F5F5F5",
        "table_alt_bg":  "#FFFFFF",
        "accent":      "#333333",
        "border":      "#CCCCCC",
    },
    "Green Formal": {
        "font":        "Calibri, Arial, sans-serif",
        "header_bg":   "#1E5631",
        "header_text": "#FFFFFF",
        "table_header_bg": "#2D8A4E",
        "table_header_text": "#FFFFFF",
        "table_row_bg":  "#EBF5EE",
        "table_alt_bg":  "#FFFFFF",
        "accent":      "#2D8A4E",
        "border":      "#A8D5B5",
    },
    "Maroon Classic": {
        "font":        "'Times New Roman', Georgia, serif",
        "header_bg":   "#6B1A1A",
        "header_text": "#FFFFFF",
        "table_header_bg": "#9B2335",
        "table_header_text": "#FFFFFF",
        "table_row_bg":  "#FAF0F0",
        "table_alt_bg":  "#FFFFFF",
        "accent":      "#9B2335",
        "border":      "#E8C5C5",
    },
    "Clean Minimal": {
        "font":        "Arial, Helvetica, sans-serif",
        "header_bg":   "#F8F8F8",
        "header_text": "#222222",
        "table_header_bg": "#EEEEEE",
        "table_header_text": "#222222",
        "table_row_bg":  "#FAFAFA",
        "table_alt_bg":  "#FFFFFF",
        "accent":      "#555555",
        "border":      "#DDDDDD",
    },
}

def build_html(plain_body, row, cfg, style):
    """Convert plain-text body to a well-structured HTML email."""
    s = EMAIL_STYLES[style]
    font     = s["font"]
    h_bg     = s["header_bg"]
    h_txt    = s["header_text"]
    th_bg    = s["table_header_bg"]
    th_txt   = s["table_header_text"]
    tr_bg    = s["table_row_bg"]
    ta_bg    = s["table_alt_bg"]
    accent   = s["accent"]
    border   = s["border"]

    company  = cfg.get("company", "")
    signatory= cfg.get("signatory", "")
    reply_to = cfg.get("reply_to", "")

    contact  = str(row.get("Contact Person") or "Sir/Madam")
    party    = str(row.get("Party Name", ""))
    amount   = fmt_amount(row.get("Outstanding Balance", 0), row.get("Currency") or "INR")
    ref      = str(row.get("Reference / Invoice No.") or "N/A")
    due      = str(row.get("Due Date", ""))[:10] if row.get("Due Date") else "as agreed"
    remarks  = str(row.get("Remarks") or "—")
    conf_dt  = cfg.get("conf_date", "")
    flow     = "payable to your organisation" if "AP" in str(row.get("Type (AR/AP)", "")).upper() else "receivable from your organisation"

    table_rows = [
        ("Party Name",         party),
        ("Outstanding Amount", f"<strong>{amount}</strong>"),
        ("Reference",          ref),
        ("Due Date",           due),
        ("Remarks",            remarks),
    ]
    tr_html = ""
    for idx, (label, value) in enumerate(table_rows):
        bg = tr_bg if idx % 2 == 0 else ta_bg
        tr_html += f"""
        <tr>
          <td style="padding:9px 14px;font-weight:600;color:#444;background:{bg};border-bottom:1px solid {border};width:38%;">{label}</td>
          <td style="padding:9px 14px;background:{bg};border-bottom:1px solid {border};">{value}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:{font};">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f4f4;padding:30px 0;">
 <tr><td align="center">
  <table width="620" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:6px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.08);">

   <!-- Header -->
   <tr>
     <td style="background:{h_bg};padding:22px 30px;">
       <div style="font-size:18px;font-weight:700;color:{h_txt};letter-spacing:0.3px;">{company}</div>
       <div style="font-size:12px;color:{h_txt};opacity:0.8;margin-top:3px;">Balance Confirmation · {conf_dt}</div>
     </td>
   </tr>

   <!-- Body -->
   <tr>
     <td style="padding:28px 30px;font-size:14px;line-height:1.7;color:#333;">
       <p style="margin:0 0 16px;">Dear {contact},</p>
       <p style="margin:0 0 16px;">Greetings from <strong>{company}</strong>!</p>
       <p style="margin:0 0 16px;">As part of our periodic balance confirmation exercise, we kindly request you to confirm the outstanding balance as on <strong>{conf_dt}</strong>.</p>
       <p style="margin:0 0 20px;">As per our records, the following amount is <strong>{flow}</strong>:</p>

       <!-- Balance Table -->
       <table width="100%" cellpadding="0" cellspacing="0" style="border-radius:4px;overflow:hidden;border:1px solid {border};margin-bottom:24px;font-size:13px;">
         <tr>
           <th colspan="2" style="background:{th_bg};color:{th_txt};padding:10px 14px;text-align:left;font-size:12px;letter-spacing:0.5px;text-transform:uppercase;">Balance Details</th>
         </tr>
         {tr_html}
       </table>

       <p style="margin:0 0 16px;">Kindly confirm the above balance by replying to this email. If there is any discrepancy, please share the relevant details and supporting documents so that we may reconcile at the earliest.</p>
       <p style="margin:0 0 16px;">Your prompt response will be greatly appreciated.</p>
       <p style="margin:0 0 4px;">For any queries, please write to us at: <a href="mailto:{reply_to}" style="color:{accent};">{reply_to}</a></p>
     </td>
   </tr>

   <!-- Signature -->
   <tr>
     <td style="padding:0 30px 28px;font-size:14px;color:#333;">
       <p style="margin:0 0 4px;">Warm regards,</p>
       <p style="margin:0 0 2px;font-weight:700;color:{accent};">{signatory}</p>
       <p style="margin:0;color:#666;font-size:13px;">{company}</p>
     </td>
   </tr>

   <!-- Footer -->
   <tr>
     <td style="background:{h_bg};padding:12px 30px;text-align:center;">
       <p style="margin:0;font-size:11px;color:{h_txt};opacity:0.7;">This is a system-generated balance confirmation email. Please do not reply to this footer.</p>
     </td>
   </tr>

  </table>
 </td></tr>
</table>
</body></html>"""
    return html

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def fmt_amount(value, currency="INR"):
    try: return f"{currency} {float(value):,.2f}"
    except: return f"{currency} 0.00"

def extract_email(raw):
    raw = str(raw or "").strip()
    m = re.search(r"<([^@]+@[^@]+\.[^@]+)>", raw)
    return m.group(1).strip() if m else raw

def valid_email(e):
    return bool(e and re.match(r"[^@]+@[^@]+\.[^@]+", extract_email(str(e))))

def merge_cc(row_cc, global_cc):
    parts = []
    for raw in (row_cc, global_cc):
        for addr in re.split(r"[,;]", str(raw or "")):
            addr = extract_email(addr)
            if addr and addr not in parts:
                parts.append(addr)
    return ", ".join(parts)

def resolve(template, row, cfg):
    result = template
    for ph, fn in PLACEHOLDERS.items():
        result = result.replace(ph, fn(row, cfg))
    return result

def find_header_row(ws):
    for n in (1, 2):
        vals = [str(c.value).strip() if c.value else "" for c in ws[n]]
        if "Party Name" in vals and "Email ID" in vals:
            return n
    return 2

def read_excel(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Party List" not in wb.sheetnames:
        return None, "Sheet 'Party List' not found."
    ws = wb["Party List"]
    hr = find_header_row(ws)
    headers = [str(c.value).strip() if c.value else "" for c in ws[hr]]
    rows = []
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        if not any(row): continue
        rec = dict(zip(headers, row))
        if rec.get("Party Name") and rec.get("Email ID"):
            rec["Email ID"] = extract_email(rec["Email ID"])
            rows.append(rec)
    return rows, None

# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════

_defaults = {
    "rows": [], "drafts": [], "send_log": [], "file_id": None,
    "subject_tpl": DEFAULT_SUBJECT, "body_tpl": DEFAULT_BODY,
    "cfg_company":     "",
    "cfg_signatory":   "",
    "cfg_reply_to":    "",
    "cfg_conf_date":   datetime.today().date(),
    "cfg_global_cc":   "",
    "cfg_type_filter": ["AR", "AP"],
    "cfg_provider":    "Desktop Outlook App (win32com)",
    "cfg_action":      "Save as Drafts",
    "cfg_smtp_user":   "",
    "cfg_smtp_pass":   "",
    "cfg_from_name":   "",
    "cfg_style":       "Professional Blue",
}
for k, v in _defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ══════════════════════════════════════════════════════════════════════════════
# HEADER
# ══════════════════════════════════════════════════════════════════════════════

hdr_col, help_col = st.columns([11, 1])
with hdr_col:
    st.markdown("""
<div style="background:var(--secondary-background-color);padding:24px 32px 20px;border-radius:8px;border-bottom:4px solid #c8873a;margin-bottom:24px;">
  <div style="font-family:'IBM Plex Mono',monospace;font-size:10px;letter-spacing:2px;opacity:.7;text-transform:uppercase;margin-bottom:4px;">Financial Operations</div>
  <div style="font-size:24px;font-weight:700;">Balance Confirmation Mailer</div>
</div>""", unsafe_allow_html=True)
with help_col:
    with st.popover("❓", help="Placeholder reference"):
        st.markdown("### Placeholder Reference")
        st.markdown("""
| Placeholder | Column / Source | Default |
|---|---|---|
| `{{Party Name}}` | Party Name | — |
| `{{Email ID}}` | Email ID | — |
| `{{Contact Person}}` | Contact Person | Sir/Madam |
| `{{Type (AR/AP)}}` | Type (AR/AP) | AR |
| `{{Currency}}` | Currency | INR |
| `{{Outstanding Balance}}` | Outstanding Balance | 0.00 |
| `{{Due Date}}` | Due Date | as agreed |
| `{{Reference / Invoice No.}}` | Reference / Invoice No. | N/A |
| `{{Remarks}}` | Remarks | *(blank)* |
| `{{Confirmation Date}}` | Config | — |
| `{{Company Name}}` | Config | — |
| `{{Signatory}}` | Config | — |
| `{{Reply-To Email}}` | Config | — |
| `{{Flow}}` | Derived | receivable/payable |
""")

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

st.markdown('<div class="sec-head">⚙️ Configuration</div>', unsafe_allow_html=True)
col1, col2, col3 = st.columns(3)

with col1:
    st.markdown("**Company Details**")
    st.text_input("Company Name",            key="cfg_company",   placeholder="Horizon Industries")
    st.text_input("Signatory & Designation", key="cfg_signatory", placeholder="Ramesh Verma, CFO")
    st.text_input("Reply-To Email",          key="cfg_reply_to",  placeholder="accounts@yourcompany.com")
    st.date_input("Confirmation Date",       key="cfg_conf_date")

with col2:
    st.markdown("**Email Engine**")
    st.selectbox("Email Provider", [
        "Desktop Outlook App (win32com)",
        "Gmail (SMTP)",
        "Office 365 / Outlook (SMTP)",
    ], key="cfg_provider")
    if "Desktop" in st.session_state.cfg_provider:
        st.radio("Action", ["Save as Drafts", "Send Immediately"], key="cfg_action")
    else:
        st.info("💡 SMTP sends immediately.")
        st.text_input("Sender Email", placeholder="you@domain.com", key="cfg_smtp_user")
        st.text_input("App Password", type="password",              key="cfg_smtp_pass")
        st.text_input("Display Name", placeholder="Accounts Team",  key="cfg_from_name")

with col3:
    st.markdown("**Filters, CC & Style**")
    st.text_input("Global CC (comma-separated)", key="cfg_global_cc",
                  placeholder="manager@co.com",
                  help="Added to every email. Per-party CC can also be set in the Excel 'CC' column.")
    st.multiselect("Include types", ["AR", "AP"], key="cfg_type_filter")
    st.selectbox("Email Style / Font", list(EMAIL_STYLES.keys()), key="cfg_style",
                 help="Controls font, colours and layout of the HTML email sent to parties.")

st.divider()

cfg = {
    "company":   st.session_state.cfg_company,
    "signatory": st.session_state.cfg_signatory,
    "reply_to":  st.session_state.cfg_reply_to,
    "conf_date": st.session_state.cfg_conf_date.strftime("%d %B %Y")
                 if isinstance(st.session_state.cfg_conf_date, (datetime, date))
                 else str(st.session_state.cfg_conf_date),
}
global_cc   = st.session_state.cfg_global_cc
type_filter = st.session_state.cfg_type_filter
email_style = st.session_state.cfg_style

# ══════════════════════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════════════════════

tab1, tab2, tab3, tab4 = st.tabs([
    "📂 Step 1 · Upload & Edit",
    "✍️ Step 2 · Template",
    "✏️ Step 3 · Review",
    "📤 Step 4 · Dispatch",
])

# ── TAB 1 — UPLOAD ────────────────────────────────────────────────────────────
with tab1:
    uploaded = st.file_uploader("Upload Excel File", type=["xlsx", "xls"],
                                 help='Must contain sheet "Party List"')
    if uploaded and st.session_state.file_id != uploaded.file_id:
        rows, err = read_excel(uploaded.read())
        if err:
            st.error(f"❌ {err}")
        else:
            st.session_state.rows    = rows
            st.session_state.drafts  = []
            st.session_state.file_id = uploaded.file_id
            st.success(f"✅ Loaded **{len(rows)}** parties.")

    if st.session_state.rows:
        st.markdown('<div class="sec-head">Review & Edit Data</div>', unsafe_allow_html=True)
        st.caption("💡 Double-click any cell to edit before generating emails.")
        df     = pd.DataFrame(st.session_state.rows)
        edited = st.data_editor(df, use_container_width=True, num_rows="dynamic", key="data_editor")
        st.session_state.rows = edited.to_dict("records")
        ar = [r for r in st.session_state.rows if "AP" not in str(r.get("Type (AR/AP)", "")).upper()]
        ap = [r for r in st.session_state.rows if "AP"     in str(r.get("Type (AR/AP)", "")).upper()]
        c1, c2, c3 = st.columns(3)
        c1.metric("Total", len(st.session_state.rows))
        c2.metric("AR",    len(ar))
        c3.metric("AP",    len(ap))

# ── TAB 2 — TEMPLATE ──────────────────────────────────────────────────────────
with tab2:
    with st.expander("📋 Available Placeholders — click to expand", expanded=False):
        st.markdown("""
| Placeholder | Reads from | Default |
|---|---|---|
| `{{Party Name}}` | Party Name column | — |
| `{{Contact Person}}` | Contact Person column | Sir/Madam |
| `{{Outstanding Balance}}` | Outstanding Balance — formatted *CCY X,XXX.XX* | 0.00 |
| `{{Due Date}}` | Due Date column | as agreed |
| `{{Reference / Invoice No.}}` | Reference / Invoice No. column | N/A |
| `{{Remarks}}` | Remarks column | *(blank)* |
| `{{Confirmation Date}}` | Confirmation Date in config | — |
| `{{Company Name}}` | Company Name in config | — |
| `{{Signatory}}` | Signatory in config | — |
| `{{Reply-To Email}}` | Reply-To Email in config | — |
| `{{Flow}}` | Derived from Type (AR/AP) | receivable from / payable to |
""")
        st.info("💡 The body text you write here is used to build the HTML email. The balance detail table is always rendered automatically from your data — you don't need to include it manually.")

    st.text_input("Subject Template", key="subject_tpl")
    st.text_area("Body Template", key="body_tpl", height=320,
                 help="Write the email body using plain text and placeholders. The balance table and styling are applied automatically.")

    # Live style preview
    if st.session_state.cfg_company:
        with st.expander("👁️ Style Preview", expanded=False):
            sample_row = {"Party Name": "Sample Party Ltd", "Contact Person": "Mr. Sample",
                          "Type (AR/AP)": "AR", "Currency": "INR", "Outstanding Balance": 1234567.89,
                          "Due Date": "31-Mar-2026", "Reference / Invoice No.": "INV-001", "Remarks": "Sample remark"}
            st.components.v1.html(build_html("", sample_row, cfg, email_style), height=600, scrolling=True)

    if st.button("🔨 Generate Emails", type="primary", use_container_width=True):
        if not st.session_state.rows:
            st.warning("Upload a party list first (Step 1).")
        elif not cfg["company"]:
            st.warning("⚠️ Enter at least a Company Name in the Configuration above.")
        else:
            drafts = []
            for row in st.session_state.rows:
                type_tag = "AP" if "AP" in str(row.get("Type (AR/AP)", "")).upper() else "AR"
                if type_tag in type_filter and valid_email(row.get("Email ID", "")):
                    html_body = build_html(
                        resolve(st.session_state.body_tpl, row, cfg),
                        row, cfg, email_style
                    )
                    drafts.append({
                        "party":   row.get("Party Name", ""),
                        "type":    type_tag,
                        "to":      str(row.get("Email ID", "")).strip(),
                        "cc":      merge_cc(row.get("CC", ""), global_cc),
                        "subject": resolve(st.session_state.subject_tpl, row, cfg),
                        "body":    html_body,
                        "include": True,
                        "row":     row,
                    })
            for i, d in enumerate(drafts):
                st.session_state[f"rv_inc_{i}"] = d["include"]
                st.session_state[f"rv_to_{i}"]  = d["to"]
                st.session_state[f"rv_cc_{i}"]  = d["cc"]
                st.session_state[f"rv_s_{i}"]   = d["subject"]
                st.session_state[f"rv_b_{i}"]   = d["body"]
            st.session_state.drafts = drafts
            st.success(f"✅ Generated {len(drafts)} emails — go to Step 3 to review.")

# ── TAB 3 — REVIEW ────────────────────────────────────────────────────────────
with tab3:
    if not st.session_state.drafts:
        st.info("Generate emails in Step 2 first.")
    else:
        for i, draft in enumerate(st.session_state.drafts):
            with st.expander(f"✉️ {draft['party']} | {draft['type']}", expanded=False):
                st.checkbox("Include in dispatch", key=f"rv_inc_{i}")
                st.text_input("To",      key=f"rv_to_{i}")
                st.text_input("CC",      key=f"rv_cc_{i}")
                st.text_input("Subject", key=f"rv_s_{i}")
                st.markdown("**Email Preview**")
                st.components.v1.html(st.session_state.get(f"rv_b_{i}", ""), height=520, scrolling=True)

        for i, draft in enumerate(st.session_state.drafts):
            draft["include"] = st.session_state.get(f"rv_inc_{i}", True)
            draft["to"]      = st.session_state.get(f"rv_to_{i}",  draft["to"])
            draft["cc"]      = st.session_state.get(f"rv_cc_{i}",  draft["cc"])
            draft["subject"] = st.session_state.get(f"rv_s_{i}",   draft["subject"])
            draft["body"]    = st.session_state.get(f"rv_b_{i}",   draft["body"])

# ── TAB 4 — DISPATCH ──────────────────────────────────────────────────────────
with tab4:
    provider  = st.session_state.cfg_provider
    action    = st.session_state.cfg_action
    smtp_user = st.session_state.cfg_smtp_user
    smtp_pass = st.session_state.cfg_smtp_pass
    from_name = st.session_state.cfg_from_name
    selected  = [d for d in st.session_state.drafts if d.get("include")]

    if not selected:
        st.warning("No emails selected. Go to Step 3 and check the boxes.")
    else:
        st.write(f"**Ready:** {len(selected)} emails via `{provider}`")

        if st.button(f"🚀 {action} ({len(selected)} emails)", type="primary", use_container_width=True):
            if "SMTP" in provider and (not smtp_user or not smtp_pass):
                st.error("❌ Enter Sender Email and App Password in Configuration.")
                st.stop()

            progress = st.progress(0)
            status   = st.empty()
            results  = []
            t0       = time.time()
            server = outlook = None

            try:
                if "Desktop" in provider:
                    import win32com.client, pythoncom
                    pythoncom.CoInitialize()
                    outlook = win32com.client.Dispatch("Outlook.Application")
                elif "Gmail" in provider:
                    ctx    = ssl.create_default_context()
                    server = smtplib.SMTP_SSL("smtp.gmail.com", 465, context=ctx)
                    server.login(smtp_user, smtp_pass)
                else:
                    server = smtplib.SMTP("smtp.office365.com", 587)
                    server.starttls()
                    server.login(smtp_user, smtp_pass)
            except Exception as e:
                st.error(f"Connection error: {e}")
                st.stop()

            for idx, d in enumerate(selected):
                status.markdown(f"**Sending:** {d['party']} ({idx+1}/{len(selected)})")
                try:
                    if "Desktop" in provider:
                        mail            = outlook.CreateItem(0)
                        mail.To         = d["to"]
                        mail.CC         = d.get("cc", "")
                        mail.Subject    = d["subject"]
                        mail.HTMLBody   = d["body"]   # HTML email
                        if "Send" in action:
                            mail.Send()
                            results.append({"Party": d["party"], "To": d["to"], "CC": d.get("cc",""), "Status": "✅ Sent"})
                        else:
                            mail.Save()
                            results.append({"Party": d["party"], "To": d["to"], "CC": d.get("cc",""), "Status": "✅ Draft Saved"})
                    else:
                        msg            = MIMEMultipart("alternative")
                        msg["Subject"] = d["subject"]
                        msg["From"]    = f"{from_name or cfg['company']} <{smtp_user}>"
                        msg["To"]      = d["to"]
                        cc_list        = [a.strip() for a in d.get("cc","").split(",") if a.strip()]
                        if cc_list: msg["Cc"] = ", ".join(cc_list)
                        msg.attach(MIMEText(d["body"], "html"))   # HTML email
                        server.sendmail(smtp_user, [d["to"]] + cc_list, msg.as_string())
                        results.append({"Party": d["party"], "To": d["to"], "CC": d.get("cc",""), "Status": "✅ Sent"})
                        if idx < len(selected) - 1:
                            time.sleep(0.5)
                except Exception as e:
                    results.append({"Party": d["party"], "To": d["to"], "CC": d.get("cc",""), "Status": f"❌ {e}"})

                progress.progress((idx + 1) / len(selected))

            if server:
                try: server.quit()
                except: pass

            st.session_state.send_log = results
            status.empty()
            progress.empty()

            ok   = sum(1 for r in results if "✅" in r["Status"])
            fail = sum(1 for r in results if "❌" in r["Status"])
            secs = round(time.time() - t0, 1)
            st.markdown(f"""
            <div style="display:flex;gap:15px;margin-bottom:20px;">
              <div class="metric-box" style="flex:1;"><div style="font-size:12px;opacity:0.7;">✅ Success</div><div class="metric-val" style="color:#2e7d32;">{ok}</div></div>
              <div class="metric-box" style="flex:1;"><div style="font-size:12px;opacity:0.7;">❌ Failed</div><div class="metric-val" style="color:#d32f2f;">{fail}</div></div>
              <div class="metric-box" style="flex:1;"><div style="font-size:12px;opacity:0.7;">⏱️ Time</div><div class="metric-val">{secs}s</div></div>
            </div>""", unsafe_allow_html=True)

    if st.session_state.send_log:
        st.markdown('<div class="sec-head">Dispatch Log</div>', unsafe_allow_html=True)
        st.dataframe(pd.DataFrame(st.session_state.send_log), use_container_width=True)
